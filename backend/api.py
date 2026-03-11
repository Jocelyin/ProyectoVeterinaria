from datetime import timedelta, datetime
from typing import List, Optional
from fastapi import FastAPI, Depends, HTTPException, status, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from pydantic import BaseModel
import io

from database import engine, get_db, Base
import models
import schemas
from notifications import notify_cita_created
from auth import (
    authenticate_user,
    create_access_token,
    get_current_user,
    get_password_hash,
    ACCESS_TOKEN_EXPIRE_MINUTES,
)

# Crear las tablas en la base de datos
Base.metadata.create_all(bind=engine)

app = FastAPI(title="Veterinaria Test - API Multi-Tenant", version="3.1.0")

# Configurar CORS — permite cualquier puerto de localhost para desarrollo local
app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=r"http://(localhost|127\.0\.0\.1)(:\d+)?",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def create_initial_admin(db: Session):
    """Crear usuario admin si no existe"""
    admin = db.query(models.User).filter(models.User.username == "admin").first()
    if not admin:
        admin_user = models.User(
            username="admin",
            hashed_password=get_password_hash("admin"),
            nombre_veterinaria="Sistema Admin",
            direccion="Oficina Central",
            telefono="000-000-0000",
            email="admin@veterinariatest.com",
            role="admin",
            is_active=True
        )
        db.add(admin_user)
        db.commit()
        print("[OK] Usuario admin/admin creado exitosamente")


@app.on_event("startup")
def startup_event():
    db = next(get_db())
    create_initial_admin(db)
    db.close()


@app.post("/api/login", response_model=schemas.Token)
def login(login_data: schemas.UserLogin, db: Session = Depends(get_db)):
    user = authenticate_user(db, login_data.username, login_data.password)
    # Autenticación más robusta (revisar auth.py en futuro si se necesita)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Usuario o contraseña incorrectos",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    # Incluir Rol en el token si se desea, pero por ahora solo username
    access_token = create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}


@app.post("/api/register", response_model=schemas.UserResponse)
def register_veterinaria(user_data: schemas.UserRegister, db: Session = Depends(get_db)):
    """Registro de nuevas veterinarias"""
    if db.query(models.User).filter(models.User.username == user_data.username).first():
        raise HTTPException(status_code=400, detail="El usuario ya existe")
    
    if db.query(models.User).filter(models.User.email == user_data.email).first():
        raise HTTPException(status_code=400, detail="El email ya está registrado")
    
    new_user = models.User(
        username=user_data.username,
        hashed_password=get_password_hash(user_data.password),
        nombre_veterinaria=user_data.nombre_veterinaria,
        direccion=user_data.direccion,
        telefono=user_data.telefono,
        email=user_data.email,
        role="veterinaria",  # Por defecto
        is_active=True
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user


@app.get("/api/me", response_model=schemas.UserResponse)
def get_current_user_info(current_user: models.User = Depends(get_current_user)):
    return current_user


@app.put("/api/me", response_model=schemas.UserResponse)
def update_current_user(
    user_update: schemas.UserUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    update_data = user_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(current_user, field, value)
    db.commit()
    db.refresh(current_user)
    return current_user


# === PACIENTES (CLIENTE + MASCOTA) ===

@app.get("/api/patients", response_model=List[schemas.PatientResponse])
def get_my_patients(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Obtener listado 'flattened' de mascotas con sus dueños (para el dashboard)"""
    if current_user.role == "admin":
        # Admin podría ver todo o nada, por ahora nada o dashboard admin
        raise HTTPException(status_code=403, detail="Los administradores deben usar el panel de administración")
        
    # Query mascotas JOIN clientes WHERE cliente.veterinaria_id == current_user.id
    results = (
        db.query(models.Mascota)
        .join(models.Cliente)
        .filter(models.Cliente.veterinaria_id == current_user.id)
        .options(joinedload(models.Mascota.cliente))
        .all()
    )
    
    # Map to schema
    response_data = []
    for mascota in results:
        response_data.append({
            "id": mascota.id,
            "nombre_mascota": mascota.nombre,
            "especie": mascota.especie,
            "raza": mascota.raza,
            "edad": mascota.edad,
            "peso": mascota.peso,
            "sexo": mascota.sexo,
            "datos_extra": mascota.datos_extra,
            "created_at": None, # Mascota doesn't have created_at, using Cliente's? Or just null
            "cliente_id": mascota.cliente_id,
            "nombre_propietario": mascota.cliente.nombre,
            "telefono_propietario": mascota.cliente.telefono
        })
        
    return response_data


@app.post("/api/patients", response_model=schemas.PatientResponse)
def create_patient(
    patient_data: schemas.PatientCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Crear Cliente y Mascota en un solo paso"""
    if current_user.role != "veterinaria":
        raise HTTPException(status_code=403, detail="Solo las veterinarias pueden registrar pacientes")
        
    try:
        # 1. Crear Cliente
        new_client = models.Cliente(
            nombre=patient_data.nombre_propietario,
            telefono=patient_data.telefono_propietario,
            veterinaria_id=current_user.id,
            direccion="" # Opcional
        )
        db.add(new_client)
        db.flush() # Para obtener ID
        
        # 2. Crear Mascota
        new_mascota = models.Mascota(
            nombre=patient_data.nombre_mascota,
            especie=patient_data.especie,
            raza=patient_data.raza,
            edad=patient_data.edad,
            datos_extra=patient_data.datos_extra,
            cliente_id=new_client.id
        )
        db.add(new_mascota)
        db.commit()
        db.refresh(new_mascota)
        
        return {
            "id": new_mascota.id,
            "nombre_mascota": new_mascota.nombre,
            "especie": new_mascota.especie,
            "raza": new_mascota.raza,
            "edad": new_mascota.edad,
            "datos_extra": new_mascota.datos_extra,
            "created_at": None,
            "cliente_id": new_client.id,
            "nombre_propietario": new_client.nombre,
            "telefono_propietario": new_client.telefono
        }
    except Exception as e:
        db.rollback()
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# === HISTORIA MEDICA ===

@app.post("/api/medical-history", response_model=schemas.HistoriaMedicaResponse)
def create_medical_history(
    history_data: schemas.HistoriaMedicaCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Crear registro medico para una mascota"""
    # Verificar que la mascota pretenece a un cliente de esta veterinaria
    mascota = db.query(models.Mascota).join(models.Cliente).filter(
        models.Mascota.id == history_data.mascota_id,
        models.Cliente.veterinaria_id == current_user.id
    ).first()
    
    if not mascota:
         raise HTTPException(status_code=404, detail="Mascota no encontrada o no pertenece a esta veterinaria")

    new_history = models.HistoriaMedica(**history_data.dict())
    db.add(new_history)
    db.commit()
    db.refresh(new_history)
    return new_history


@app.get("/api/mascotas/{mascota_id}/history", response_model=List[schemas.HistoriaMedicaResponse])
def get_medical_history(
    mascota_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    mascota = db.query(models.Mascota).join(models.Cliente).filter(
        models.Mascota.id == mascota_id,
        models.Cliente.veterinaria_id == current_user.id
    ).first()
    
    if not mascota:
         raise HTTPException(status_code=404, detail="Mascota no encontrada")

    return db.query(models.HistoriaMedica).filter(models.HistoriaMedica.mascota_id == mascota_id).order_by(models.HistoriaMedica.fecha.desc()).all()


# === MASCOTAS CRUD ===

@app.put("/api/mascotas/{mascota_id}", response_model=schemas.MascotaResponse)
def update_mascota(
    mascota_id: int,
    mascota_update: schemas.MascotaUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    mascota = db.query(models.Mascota).join(models.Cliente).filter(
        models.Mascota.id == mascota_id,
        models.Cliente.veterinaria_id == current_user.id
    ).options(joinedload(models.Mascota.cliente)).first()
    
    if not mascota:
        raise HTTPException(status_code=404, detail="Mascota no encontrada")
        
    for var, value in vars(mascota_update).items():
        if value is not None:
            setattr(mascota, var, value)
            
    db.commit()
    db.refresh(mascota)
    return {
        "id": mascota.id,
        "nombre": mascota.nombre,
        "especie": mascota.especie,
        "raza": mascota.raza,
        "edad": mascota.edad,
        "peso": mascota.peso,
        "datos_extra": mascota.datos_extra,
        "cliente_id": mascota.cliente_id,
        "cliente_nombre": mascota.cliente.nombre if mascota.cliente else None,
        "cliente_telefono": mascota.cliente.telefono if mascota.cliente else None,
    }

@app.delete("/api/mascotas/{mascota_id}")
def delete_mascota(
    mascota_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    mascota = db.query(models.Mascota).join(models.Cliente).filter(
        models.Mascota.id == mascota_id,
        models.Cliente.veterinaria_id == current_user.id
    ).first()
    
    if not mascota:
        raise HTTPException(status_code=404, detail="Mascota no encontrada")
        
    db.delete(mascota)
    db.commit()
    return {"message": "Mascota eliminada"}


# === CLIENTES CRUD ===

@app.get("/api/clients", response_model=List[schemas.ClienteResponse])
def get_my_clients(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Obtener clientes de la veterinaria autenticada con sus mascotas"""
    if current_user.role == "admin":
        raise HTTPException(status_code=403, detail="Los administradores deben usar el panel de administración")
        
    return (
        db.query(models.Cliente)
        .filter(models.Cliente.veterinaria_id == current_user.id)
        .options(joinedload(models.Cliente.mascotas))
        .all()
    )


@app.get("/api/clients/{client_id}", response_model=schemas.ClienteResponse)
def get_client_detail(
    client_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Obtener un cliente con todas sus mascotas"""
    client = (
        db.query(models.Cliente)
        .filter(models.Cliente.id == client_id, models.Cliente.veterinaria_id == current_user.id)
        .options(joinedload(models.Cliente.mascotas))
        .first()
    )
    if not client:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return client

@app.post("/api/clients", response_model=schemas.ClienteResponse)
def create_client(
    client_data: schemas.ClienteCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
     new_client = models.Cliente(
        **client_data.dict(),
        veterinaria_id=current_user.id
    )
     db.add(new_client)
     db.commit()
     db.refresh(new_client)
     return new_client


@app.post("/api/clients/{client_id}/mascotas", response_model=schemas.MascotaResponse)
def create_mascota_for_client(
    client_id: int,
    mascota_data: schemas.MascotaBase,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Crear mascota vinculada a un cliente existente"""
    client = db.query(models.Cliente).filter(
        models.Cliente.id == client_id,
        models.Cliente.veterinaria_id == current_user.id
    ).first()
    if not client:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    new_mascota = models.Mascota(
        nombre=mascota_data.nombre,
        especie=mascota_data.especie,
        raza=mascota_data.raza,
        edad=mascota_data.edad,
        peso=mascota_data.peso,
        datos_extra=mascota_data.datos_extra,
        cliente_id=client.id
    )
    db.add(new_mascota)
    db.commit()
    db.refresh(new_mascota)
    return {
        "id": new_mascota.id,
        "nombre": new_mascota.nombre,
        "especie": new_mascota.especie,
        "raza": new_mascota.raza,
        "edad": new_mascota.edad,
        "peso": new_mascota.peso,
        "datos_extra": new_mascota.datos_extra,
        "cliente_id": client.id,
        "cliente_nombre": client.nombre,
        "cliente_telefono": client.telefono,
    }

@app.put("/api/clients/{client_id}", response_model=schemas.ClienteResponse)
def update_client(
    client_id: int,
    client_update: schemas.ClienteUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    client = db.query(models.Cliente).filter(
        models.Cliente.id == client_id,
        models.Cliente.veterinaria_id == current_user.id
    ).first()
    
    if not client:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
        
    for var, value in vars(client_update).items():
        if value is not None:
            setattr(client, var, value)
            
    db.commit()
    db.refresh(client)
    return client

@app.delete("/api/clients/{client_id}")
def delete_client(
    client_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    client = db.query(models.Cliente).filter(
        models.Cliente.id == client_id,
        models.Cliente.veterinaria_id == current_user.id
    ).first()
    
    if not client:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    
    # Optional: Check if has pets/appointments before delete? Cascade rules in models handle it usually.
    db.delete(client)
    db.commit()
    return {"message": "Cliente eliminado"}


# === CITAS CRUD ===

@app.get("/api/citas", response_model=List[schemas.CitaResponse])
def get_citas(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # Get appointments for this vet (via cliente relationship usually, but Cita has direct FKs too?)
    # Model Cita has veterinaria_id. Let's check model.
    # Cita: cliente_id, mascota_id, veterinaria_id. Safe to query by veterinaria_id.
    
    citas = (
        db.query(models.Cita)
        .filter(models.Cita.veterinaria_id == current_user.id)
        .options(joinedload(models.Cita.cliente), joinedload(models.Cita.mascota))
        .all()
    )
    
    # Flatten
    response = []
    for c in citas:
        response.append({
            "id": c.id,
            "fecha_hora": c.fecha_hora,
            "motivo": c.motivo,
            "estado": c.estado,
            "cliente_id": c.cliente_id,
            "mascota_id": c.mascota_id,
            "veterinaria_id": c.veterinaria_id,
            "cliente_nombre": c.cliente.nombre if c.cliente else "Unknown",
            "mascota_nombre": c.mascota.nombre if c.mascota else "Unknown"
        })
    return response

@app.post("/api/citas", response_model=schemas.CitaResponse)
def create_cita(
    cita_data: schemas.CitaCreate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # Verify client belongs to vet
    client = db.query(models.Cliente).filter(
        models.Cliente.id == cita_data.cliente_id,
        models.Cliente.veterinaria_id == current_user.id
    ).first()
    
    if not client:
        raise HTTPException(status_code=400, detail="Cliente invalido o no pertenece a su veterinaria")
        
    new_cita = models.Cita(
        fecha_hora=cita_data.fecha_hora,
        motivo=cita_data.motivo,
        cliente_id=cita_data.cliente_id,
        mascota_id=cita_data.mascota_id,
        veterinaria_id=current_user.id,
        estado="pendiente"
    )
    db.add(new_cita)
    db.commit()
    db.refresh(new_cita)

    mascota = db.query(models.Mascota).get(cita_data.mascota_id)

    # Send confirmation notifications (background thread — non-blocking)
    notify_cita_created(
        email=client.email,
        telefono=client.telefono,
        nombre_cliente=client.nombre,
        nombre_mascota=mascota.nombre if mascota else "tu mascota",
        fecha_hora=new_cita.fecha_hora,
        nombre_veterinaria=current_user.nombre_veterinaria or "Veterinaria",
    )

    return {
        "id": new_cita.id,
        "fecha_hora": new_cita.fecha_hora,
        "motivo": new_cita.motivo,
        "estado": new_cita.estado,
        "cliente_id": new_cita.cliente_id,
        "mascota_id": new_cita.mascota_id,
        "veterinaria_id": new_cita.veterinaria_id,
        "cliente_nombre": client.nombre,
        "mascota_nombre": mascota.nombre if mascota else "Desconocida",
    }

@app.put("/api/citas/{cita_id}", response_model=schemas.CitaResponse)
def update_cita(
    cita_id: int,
    cita_update: schemas.CitaUpdate,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    cita = db.query(models.Cita).filter(
        models.Cita.id == cita_id,
        models.Cita.veterinaria_id == current_user.id
    ).first()
    
    if not cita:
        raise HTTPException(status_code=404, detail="Cita no encontrada")
        
    for var, value in vars(cita_update).items():
        if value is not None:
            setattr(cita, var, value)
            
    db.commit()
    db.refresh(cita)
    
    # Load relations for response
    # We can rely on lazy loading if not detached, or re-query
    return {
         "id": cita.id,
            "fecha_hora": cita.fecha_hora,
            "motivo": cita.motivo,
            "estado": cita.estado,
            "cliente_id": cita.cliente_id,
            "mascota_id": cita.mascota_id,
            "veterinaria_id": cita.veterinaria_id,
            "cliente_nombre": cita.cliente.nombre,
            "mascota_nombre": cita.mascota.nombre 
    }

@app.delete("/api/citas/{cita_id}")
def delete_cita(
    cita_id: int,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    cita = db.query(models.Cita).filter(
        models.Cita.id == cita_id,
        models.Cita.veterinaria_id == current_user.id
    ).first()
    
    if not cita:
        raise HTTPException(status_code=404, detail="Cita no encontrada")
        
    db.delete(cita)
    db.commit()
    return {"message": "Cita eliminada"}


# === REPORTES ===

class ReportRequest(BaseModel):
    tipo: str  # "citas", "clientes", "pacientes"
    fecha_inicio: str  # "YYYY-MM-DD"
    fecha_fin: str  # "YYYY-MM-DD"
    formato: str  # "pdf" or "excel"

@app.post("/api/reports")
def generate_report(
    report_req: ReportRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Generar reporte de Citas, Clientes o Pacientes en PDF o Excel"""
    if current_user.role == "admin":
        raise HTTPException(status_code=403, detail="Los administradores no generan reportes de veterinaria")

    try:
        fecha_inicio = datetime.strptime(report_req.fecha_inicio, "%Y-%m-%d")
        fecha_fin = datetime.strptime(report_req.fecha_fin, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(status_code=400, detail="Formato de fecha inválido. Use YYYY-MM-DD")

    if fecha_inicio > fecha_fin:
        raise HTTPException(status_code=400, detail="La fecha inicio no puede ser mayor a la fecha fin")

    tipo = report_req.tipo.lower()
    formato = report_req.formato.lower()

    # --- Gather data ---
    headers = []
    rows = []
    title = ""

    if tipo == "citas":
        title = "Reporte de Citas"
        headers = ["ID", "Fecha y Hora", "Motivo", "Estado", "Cliente", "Mascota"]
        citas = (
            db.query(models.Cita)
            .filter(
                models.Cita.veterinaria_id == current_user.id,
                models.Cita.fecha_hora >= fecha_inicio,
                models.Cita.fecha_hora <= fecha_fin
            )
            .options(joinedload(models.Cita.cliente), joinedload(models.Cita.mascota))
            .order_by(models.Cita.fecha_hora)
            .all()
        )
        for c in citas:
            rows.append([
                str(c.id),
                c.fecha_hora.strftime("%d/%m/%Y %H:%M") if c.fecha_hora else "",
                c.motivo or "",
                c.estado or "",
                c.cliente.nombre if c.cliente else "—",
                c.mascota.nombre if c.mascota else "—"
            ])

    elif tipo == "clientes":
        title = "Reporte de Clientes"
        headers = ["ID", "Nombre", "Teléfono", "Email", "Dirección", "Fecha Registro", "Nº Mascotas"]
        clientes = (
            db.query(models.Cliente)
            .filter(
                models.Cliente.veterinaria_id == current_user.id,
                models.Cliente.created_at >= fecha_inicio,
                models.Cliente.created_at <= fecha_fin
            )
            .options(joinedload(models.Cliente.mascotas))
            .order_by(models.Cliente.nombre)
            .all()
        )
        for cl in clientes:
            rows.append([
                str(cl.id),
                cl.nombre or "",
                cl.telefono or "",
                cl.email or "",
                cl.direccion or "",
                cl.created_at.strftime("%d/%m/%Y") if cl.created_at else "",
                str(len(cl.mascotas)) if cl.mascotas else "0"
            ])

    elif tipo == "pacientes":
        title = "Reporte de Pacientes"
        headers = ["ID", "Nombre", "Especie", "Raza", "Edad", "Peso", "Sexo", "Propietario"]
        mascotas = (
            db.query(models.Mascota)
            .join(models.Cliente)
            .filter(
                models.Cliente.veterinaria_id == current_user.id,
                models.Cliente.created_at >= fecha_inicio,
                models.Cliente.created_at <= fecha_fin
            )
            .options(joinedload(models.Mascota.cliente))
            .order_by(models.Mascota.nombre)
            .all()
        )
        for m in mascotas:
            rows.append([
                str(m.id),
                m.nombre or "",
                m.especie or "",
                m.raza or "",
                m.edad or "",
                m.peso or "",
                m.sexo or "",
                m.cliente.nombre if m.cliente else "—"
            ])
    else:
        raise HTTPException(status_code=400, detail="Tipo de reporte inválido. Use: citas, clientes, pacientes")

    vet_name = current_user.nombre_veterinaria or "Veterinaria"
    subtitle = f"{report_req.fecha_inicio} al {report_req.fecha_fin}"

    # --- Generate file ---
    try:
        if formato == "excel":
            return _generate_excel(title, subtitle, vet_name, headers, rows, tipo)
        elif formato == "pdf":
            return _generate_pdf(title, subtitle, vet_name, headers, rows, tipo)
        else:
            raise HTTPException(status_code=400, detail="Formato inválido. Use: pdf, excel")
    except Exception as e:
        import traceback
        with open("pdf_error.txt", "w") as f:
            f.write(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


def _generate_excel(title, subtitle, vet_name, headers, rows, tipo):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = title

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"{vet_name} — {title}")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center")

    # Subtitle row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sub_cell = ws.cell(row=2, column=1, value=f"Período: {subtitle}")
    sub_cell.font = Font(size=10, italic=True, color="666666")
    sub_cell.alignment = Alignment(horizontal="center")

    # Headers
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    header_font = Font(bold=True, color="10B981", size=10)
    thin_border = Border(
        bottom=Side(style='thin', color='333333')
    )
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row_data in enumerate(rows, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center")

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_data in rows:
            if col_idx - 1 < len(row_data):
                max_len = max(max_len, len(str(row_data[col_idx - 1])))
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = min(max_len + 4, 30)

    # Total
    total_row = len(rows) + 6
    ws.cell(row=total_row, column=1, value=f"Total de registros: {len(rows)}").font = Font(bold=True, size=10)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"reporte_{tipo}_{subtitle.replace(' al ', '_')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def _generate_pdf(title, subtitle, vet_name, headers, rows, tipo):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch

    buffer = io.BytesIO()
    page_size = landscape(letter) if len(headers) > 6 else letter
    doc = SimpleDocTemplate(buffer, pagesize=page_size, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Title'],
        fontSize=18, textColor=colors.HexColor("#059669"), spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor("#6b7280"), spaceAfter=20
    )

    elements.append(Paragraph(f"{vet_name} — {title}", title_style))
    elements.append(Paragraph(f"Período: {subtitle}", subtitle_style))
    elements.append(Spacer(1, 12))

    # Table
    table_data = [headers] + rows
    table = Table(table_data, repeatRows=1)

    table_style = TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#111827")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#10B981")),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        # Body
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        # Alternating rows
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
    ])
    table.setStyle(table_style)
    elements.append(table)

    # Total
    elements.append(Spacer(1, 20))
    total_style = ParagraphStyle(
        'Total', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor("#374151")
    )
    elements.append(Paragraph(f"<b>Total de registros:</b> {len(rows)}", total_style))

    doc.build(elements)
    buffer.seek(0)

    filename = f"reporte_{tipo}_{subtitle.replace(' al ', '_')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# === ADMIN ENDPOINTS ===

@app.get("/api/admin/veterinarias", response_model=List[schemas.UserResponse])
def get_all_veterinarias(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """(Admin) Ver todas las veterinarias registradas"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Acceso denegado. Se requieren permisos de administrador")
    
    veterinarias = db.query(models.User).filter(models.User.role == "veterinaria").all()
    return veterinarias


@app.get("/api/health")
def health_check():
    return {"status": "ok", "message": "Veterinaria Multi-Tenant API V3.1 OK"}
